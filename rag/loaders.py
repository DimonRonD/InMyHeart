# -*- coding: utf-8 -*-
"""Извлечение текста из файлов source/."""
from __future__ import annotations

import csv
import io
from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


def load_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return path.read_text(encoding="utf-8")
    if suffix == ".docx":
        return _load_docx(path)
    if suffix == ".pdf":
        return _load_pdf(path)
    if suffix == ".xlsx":
        return _load_xlsx(path)
    raise ValueError(f"Неподдерживаемый формат: {path}")


def _load_docx(path: Path) -> str:
    doc = DocxDocument(path)
    parts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(parts)


def _load_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text.strip())
    return "\n\n".join(parts)


def _load_xlsx(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append("\t".join("" if c is None else str(c) for c in row))
    wb.close()
    return "\n".join(rows)


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    raw = path.read_text(encoding="utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw), delimiter=";")
    return [{k.strip(): (v or "").strip() for k, v in row.items()} for row in reader]
