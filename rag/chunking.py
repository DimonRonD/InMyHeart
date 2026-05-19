# -*- coding: utf-8 -*-
"""
Разбиение документов source/ на чанки по chunck_splitting.md.
Каждый чанк содержит metadata с путём к исходному файлу и позицией.
"""
from __future__ import annotations

import hashlib
from pathlib import Path

from rag.loaders import load_text, read_csv_rows
from rag.models import TextChunk

# FAQ: канонический CSV; TXT/MD не индексируем (дубли)
FAQ_SKIP_NAMES = {
    "FAQ_zapis_rezhim.txt",
    "FAQ_analizy_diagnostika.txt",
    "FAQ_organizaciya_servis.md",
}

SKIP_FILES = {"README.md", "raspisanie_google_sheets.md"}


def _base_meta(path: Path, source_root: Path, doc_type: str, **extra: str) -> dict:
    rel = path.relative_to(source_root).as_posix()
    meta = {
        "source_file": rel,
        "source_filename": path.name,
        "source_folder": path.parent.name,
        "doc_type": doc_type,
        "language": "ru",
    }
    meta.update({k: v for k, v in extra.items() if v})
    return meta


def _chunk_id(source_file: str, suffix: str) -> str:
    raw = f"{source_file}::{suffix}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _slug_from_stem(stem: str, prefix: str) -> str:
    if stem.startswith(prefix):
        return stem[len(prefix) :]
    return stem


def chunk_file(path: Path, source_root: Path) -> list[TextChunk]:
    name = path.name
    if name in SKIP_FILES:
        return []

    parts = path.relative_to(source_root).parts

    if "faq" in parts:
        return _chunk_faq(path, source_root)

    if "uslugi" in parts:
        return _chunk_uslugi(path, source_root)

    if "raspisanie" in parts:
        return _chunk_raspisanie(path, source_root)

    if "podgotovka_analizov" in parts:
        return _chunk_preparation(path, source_root)

    if "pamyatki_pacientov" in parts:
        return _chunk_pamyatka(path, source_root)

    if "zapis_priema" in parts:
        return _chunk_zapis(path, source_root)

    if "poseshchenie_kliniki" in parts:
        return _chunk_poseshchenie(path, source_root)

    return _chunk_whole_file(path, source_root, doc_type="other")


def chunk_source_directory(source_root: Path) -> list[TextChunk]:
    chunks: list[TextChunk] = []
    for path in sorted(source_root.rglob("*")):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".txt", ".md", ".csv", ".pdf", ".docx", ".xlsx"}:
            continue
        chunks.extend(chunk_file(path, source_root))
    return chunks


# --- FAQ ---


def _chunk_faq(path: Path, source_root: Path) -> list[TextChunk]:
    if path.name in FAQ_SKIP_NAMES:
        return []
    if path.suffix.lower() != ".csv":
        return []

    rows = read_csv_rows(path)
    rel = path.relative_to(source_root).as_posix()
    chunks = []
    for row in rows:
        q = row.get("вопрос", row.get("question", ""))
        a = row.get("ответ", row.get("answer", ""))
        cat = row.get("категория", row.get("category", ""))
        row_id = row.get("id", "")
        if not q or not a:
            continue
        content = f"Вопрос: {q}\nОтвет: {a}"
        meta = _base_meta(path, source_root, "faq", category=cat, faq_id=str(row_id))
        meta["chunk_index"] = str(row_id)
        meta["chunk_id"] = _chunk_id(rel, f"faq-{row_id}")
        chunks.append(TextChunk(content=content, metadata=meta))
    return chunks


# --- Услуги ---


def _chunk_uslugi(path: Path, source_root: Path) -> list[TextChunk]:
    suffix = path.suffix.lower()
    rel = path.relative_to(source_root).as_posix()

    if suffix == ".csv":
        return _chunk_uslugi_csv(path, source_root, rel)
    if suffix == ".xlsx":
        return _chunk_uslugi_xlsx(path, source_root, rel)
    if suffix == ".txt":
        return _chunk_uslugi_txt(path, source_root, rel)
    return []


def _chunk_uslugi_csv(path: Path, source_root: Path, rel: str) -> list[TextChunk]:
    chunks = []
    for i, row in enumerate(read_csv_rows(path), 1):
        code = row.get("код", "")
        content = (
            f"Услуга {code}: {row.get('название', '')}. "
            f"Категория: {row.get('категория', '')}. "
            f"Цена: {row.get('цена_rub', '')} ₽. "
            f"Длительность: {row.get('длительность_мин', '')} мин. "
            f"Подготовка: {row.get('подготовка', '')}. "
            f"Описание: {row.get('описание', '')}."
        )
        meta = _base_meta(
            path,
            source_root,
            "service",
            category=row.get("категория", ""),
            service_code=code,
            prep_doc=row.get("подготовка", ""),
            price=str(row.get("цена_rub", "")),
        )
        meta["chunk_index"] = str(i)
        meta["chunk_id"] = _chunk_id(rel, code or str(i))
        chunks.append(TextChunk(content=content, metadata=meta))
    return chunks


def _chunk_uslugi_xlsx(path: Path, source_root: Path, rel: str) -> list[TextChunk]:
    from openpyxl import load_workbook

    chunks = []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    for i, row in enumerate(rows[1:], 1):
        data = {headers[j]: ("" if v is None else str(v).strip()) for j, v in enumerate(row) if j < len(headers)}
        code = data.get("код", "")
        content = (
            f"Услуга {code}: {data.get('название', '')}. "
            f"Категория: {data.get('категория', '')}. "
            f"Цена: {data.get('цена_rub', '')} ₽. "
            f"Длительность: {data.get('длительность_мин', '')} мин. "
            f"Подготовка: {data.get('подготовка', '')}. "
            f"Описание: {data.get('описание', '')}."
        )
        meta = _base_meta(
            path,
            source_root,
            "service",
            category=data.get("категория", ""),
            service_code=code,
            prep_doc=data.get("подготовка", ""),
            price=str(data.get("цена_rub", "")),
        )
        meta["chunk_index"] = str(i)
        meta["chunk_id"] = _chunk_id(rel, code or str(i))
        chunks.append(TextChunk(content=content, metadata=meta))
    return chunks


def _chunk_uslugi_txt(path: Path, source_root: Path, rel: str) -> list[TextChunk]:
    lines = path.read_text(encoding="utf-8").splitlines()
    chunks = []
    idx = 0
    for line in lines:
        line = line.strip()
        if not line or line.startswith("Перечень") or line.startswith("Позиций") or line.startswith("Формат"):
            continue
        if not line.startswith("INM-"):
            continue
        idx += 1
        parts = [p.strip() for p in line.split("|")]
        code = parts[0] if parts else ""
        meta = _base_meta(path, source_root, "service", service_code=code)
        meta["chunk_index"] = str(idx)
        meta["chunk_id"] = _chunk_id(rel, code)
        chunks.append(TextChunk(content=f"Услуга (диагностика/процедуры). {line}", metadata=meta))
    return chunks


# --- Расписание ---


def _chunk_raspisanie(path: Path, source_root: Path) -> list[TextChunk]:
    name = path.name
    rel = path.relative_to(source_root).as_posix()

    if name == "raspisanie_google_sheets.md":
        return []

    if name == "raspisanie_prazdniki.txt":
        text = load_text(path)
        meta = _base_meta(path, source_root, "schedule", topic="prazdniki")
        meta["chunk_index"] = "0"
        meta["chunk_id"] = _chunk_id(rel, "whole")
        return [TextChunk(content=text, metadata=meta)]

    if name == "raspisanie_rabota_kliniki.csv":
        return _chunk_schedule_clinic_rows(path, source_root, rel)

    if name in ("raspisanie_vrachey.csv", "raspisanie_laboratoriya_diagnostika.csv"):
        return _chunk_schedule_generic_rows(path, source_root, rel)

    return _chunk_whole_file(path, source_root, doc_type="schedule")


def _chunk_schedule_clinic_rows(path: Path, source_root: Path, rel: str) -> list[TextChunk]:
    chunks = []
    for i, row in enumerate(read_csv_rows(path), 1):
        day = row.get("день_недели", "")
        content = (
            f"Режим работы клиники INMYHEART, {day}: "
            f"открытие {row.get('открытие', '')}, закрытие {row.get('закрытие', '')}. "
            f"Перерыв: {row.get('перерыв_с', '')}–{row.get('перерыв_до', '')}. "
            f"Примечание: {row.get('примечание', '')}."
        )
        meta = _base_meta(path, source_root, "schedule", topic="rabota_kliniki", day=day)
        meta["chunk_index"] = str(i)
        meta["chunk_id"] = _chunk_id(rel, day)
        chunks.append(TextChunk(content=content.strip(), metadata=meta))
    return chunks


def _chunk_schedule_generic_rows(path: Path, source_root: Path, rel: str) -> list[TextChunk]:
    chunks = []
    is_doctors = "vrachey" in path.name
    for i, row in enumerate(read_csv_rows(path), 1):
        if is_doctors:
            spec = row.get("специальность", "")
            doctor = row.get("врач", "")
            schedule = ", ".join(f"{d} {row.get(d, '')}" for d in ("пн", "вт", "ср", "чт", "пт", "сб") if row.get(d))
            content = f"{spec} {doctor}: расписание — {schedule}."
            meta = _base_meta(
                path, source_root, "schedule", topic="vrachi",
                specialty=spec, doctor=doctor,
            )
            key = doctor
        else:
            dept = row.get("отделение", "")
            content = (
                f"Расписание отделения «{dept}»: "
                f"будни {row.get('пн-пт', '')}, суббота {row.get('сб', '')}, "
                f"воскресенье {row.get('вс', '')}. {row.get('примечание', '')}"
            )
            meta = _base_meta(path, source_root, "schedule", topic="laboratoriya", department=dept)
            key = dept
        meta["chunk_index"] = str(i)
        meta["chunk_id"] = _chunk_id(rel, key or str(i))
        chunks.append(TextChunk(content=content.strip(), metadata=meta))
    return chunks


# --- Подготовка, памятки, запись, посещение ---


def _chunk_preparation(path: Path, source_root: Path) -> list[TextChunk]:
    text = load_text(path)
    slug = _slug_from_stem(path.stem, "podgotovka_")
    meta = _base_meta(path, source_root, "preparation", test_slug=slug)
    meta["chunk_index"] = "0"
    rel = path.relative_to(source_root).as_posix()
    meta["chunk_id"] = _chunk_id(rel, "whole")
    return [TextChunk(content=text, metadata=meta)]


def _chunk_pamyatka(path: Path, source_root: Path) -> list[TextChunk]:
    text = load_text(path)
    topic = _slug_from_stem(path.stem, "pamyatka_")
    meta = _base_meta(path, source_root, "pamyatka", topic=topic)
    meta["chunk_index"] = "0"
    rel = path.relative_to(source_root).as_posix()
    meta["chunk_id"] = _chunk_id(rel, "whole")
    return [TextChunk(content=text, metadata=meta)]


def _chunk_zapis(path: Path, source_root: Path) -> list[TextChunk]:
    text = load_text(path)
    stem = path.stem
    if "sposoby" in stem:
        topic = "sposoby"
    elif "otmena" in stem:
        topic = "otmena"
    elif "osobye" in stem:
        topic = "osobye_sluchai"
    else:
        topic = stem.replace("zapis_", "")
    meta = _base_meta(path, source_root, "zapis", topic=topic)
    meta["chunk_index"] = "0"
    rel = path.relative_to(source_root).as_posix()
    meta["chunk_id"] = _chunk_id(rel, "whole")
    return [TextChunk(content=text, metadata=meta)]


def _chunk_poseshchenie(path: Path, source_root: Path) -> list[TextChunk]:
    text = load_text(path)
    stem = path.stem
    if "dokumenty" in stem:
        topic = "dokumenty"
    elif "povedenie" in stem:
        topic = "povedenie"
    elif "infrastruktura" in stem:
        topic = "infrastruktura"
    else:
        topic = stem.replace("poseshchenie_", "")
    meta = _base_meta(path, source_root, "poseshchenie", topic=topic)
    meta["chunk_index"] = "0"
    rel = path.relative_to(source_root).as_posix()
    meta["chunk_id"] = _chunk_id(rel, "whole")
    return [TextChunk(content=text, metadata=meta)]


def _chunk_whole_file(path: Path, source_root: Path, doc_type: str) -> list[TextChunk]:
    text = load_text(path)
    if not text.strip():
        return []
    meta = _base_meta(path, source_root, doc_type)
    meta["chunk_index"] = "0"
    rel = path.relative_to(source_root).as_posix()
    meta["chunk_id"] = _chunk_id(rel, "whole")
    return [TextChunk(content=text, metadata=meta)]
